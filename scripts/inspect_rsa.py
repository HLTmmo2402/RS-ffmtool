# -*- coding: utf-8 -*-
"""Khảo sát cấu trúc RSA - FFM.xlsx: sheet seller + header cột 'Đơn đang đi'.
Chỉ đọc cấu trúc (KHÔNG đụng sheet credentials). Chạy: python scripts/inspect_rsa.py
"""
import sys
import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

PATH = r"d:\AI\Tool FFM\_private\RSA - FFM.xlsx"


def find_header_row(ws, max_scan=8):
    """Dò dòng header = dòng chứa 'Order ID' (hoặc 'Order')."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        vals = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[r]]
        joined = " | ".join(vals)
        if "order id" in joined or ("order" in joined and "date" in joined) or "ngày order" in joined:
            return r, vals
    return None, None


def main():
    wb = openpyxl.load_workbook(PATH, data_only=True, read_only=True)
    print("TỔNG SỐ SHEET:", len(wb.sheetnames))
    print("=" * 78)
    seller_sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        hr, _ = find_header_row(ws)
        kind = "SELLER?" if hr else "phụ"
        print("  [{:>7}] {:<28} dims={}x{}  header_row={}".format(
            kind, name[:28], ws.max_row, ws.max_column, hr))
        if hr:
            seller_sheets.append((name, hr))
    print("=" * 78)
    print("Sheet giống SELLER:", len(seller_sheets))

    # In chi tiết 1 sheet seller mẫu (đầu tiên)
    if seller_sheets:
        name, hr = seller_sheets[0]
        ws = wb[name]
        print("\n" + "=" * 78)
        print("MẪU SHEET SELLER:", name, "| header ở dòng", hr)
        print("=" * 78)
        # header
        header = [c.value for c in ws[hr]]
        print("\nCÁC CỘT (header):")
        for i, h in enumerate(header):
            if h is not None and str(h).strip():
                col_letter = openpyxl.utils.get_column_letter(i + 1)
                print("  {:>3} {:<3} {}".format(i, col_letter, str(h).strip()))
        # 3 dòng data đầu
        print("\n3 DÒNG DỮ LIỆU ĐẦU (sau header):")
        cnt = 0
        for r in range(hr + 1, ws.max_row + 1):
            row = [c.value for c in ws[r]]
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            shown = []
            for i, v in enumerate(row):
                if v is not None and str(v).strip():
                    hlabel = str(header[i]).strip() if i < len(header) and header[i] else "col%d" % i
                    shown.append("{}={}".format(hlabel[:18], str(v)[:28]))
            print("  DÒNG", r, ":")
            for s in shown[:24]:
                print("      ", s)
            cnt += 1
            if cnt >= 3:
                break

    wb.close()


if __name__ == "__main__":
    main()
